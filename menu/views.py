
from django.urls import reverse_lazy
from .models import MenuType
from .forms import MenuTypeForm
from .forms import OrganizationForm
from user.permission import IsAdminMixin, SearchMixin
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
    TemplateView,

)

from root.utils import DeleteMixin
from django.shortcuts import render


class MenuTypeMixin(IsAdminMixin):
    model = MenuType
    form_class = MenuTypeForm
    paginate_by = 50    
    queryset = MenuType.objects.all()
    success_url = reverse_lazy("menu_type_list")
    search_lookup_fields = [
        "title",
        "menutype__title",
    ]


class MenuTypeList(MenuTypeMixin, ListView):
    template_name = "scanpay/menutype/menutype-list.html"
    queryset = MenuType.objects.all()


class MenuTypeDetail(MenuTypeMixin, DetailView):
    template_name = "scanpay/menutype/menutype-list.html"


class MenuTypeCreate(MenuTypeMixin, CreateView):
    template_name = "scanpay/create.html"


class MenuTypeUpdate(MenuTypeMixin, UpdateView):
    template_name = "scanpay/update.html"


class MenuTypeDelete(MenuTypeMixin, DeleteMixin, View):
    pass


from django.views.generic import ListView
from .models import Menu, MenuType
from itertools import chain
from django.db.models import Q


    
from .forms import MenuForm


class MenuMixin(IsAdminMixin):
    model = Menu
    form_class = MenuForm
    paginate_by = 50
    queryset = Menu.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("menu_list")
    search_lookup_fields = [
        "item_name",
        "description",
    ]
    
from .models import FlagMenu

class MenuList(MenuMixin, ListView):
    template_name = "scanpay/menu/menu-list.html"
    queryset = Menu.objects.filter(status=True, is_deleted=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['flag_menu_active'] = FlagMenu.objects.first().use_same_menu_for_multiple_outlet
        context['auto_order_active'] = FlagMenu.objects.first().autoaccept_order
        context['payload_active'] = FlagMenu.objects.first().send_payload_in_notification
        context['discount_active'] = FlagMenu.objects.first().allow_discount

        return context
        
class MenuDetail(MenuMixin, DetailView):
    template_name = "scanpay/menu/menu-detail.html"


class MenuCreate(MenuMixin, CreateView):
    template_name = "scanpay/menu/menu-create.html"


class MenuUpdate(MenuMixin, UpdateView):
    template_name = "scanpay/update.html"
    

class MenuDelete(MenuMixin, DeleteMixin, View):
    pass

from django.conf import settings
from django.db import IntegrityError
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from openpyxl import load_workbook
from .models import Menu, MenuType
from urllib.parse import urlparse, urlunparse
import requests
from django.core.files.base import ContentFile

class MenuUpload(View):

    def post(self, request):
        file = request.FILES.get('file', None)
        if not file:
            messages.error(request, 'Please Provide the correct file ')
            return redirect(reverse_lazy("menu_create"))
        
        file_ext = file.name.split('.')[-1]
        if file_ext not in ['xlsx', 'xls']:
            messages.error(request, 'Format must be in xlsx or xls ')
            return redirect(reverse_lazy("menu_create"))

        wb = load_workbook(file)
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                row_data.insert(0, sheet.title)
                excel_data.append(row_data)
        
        for data in excel_data:
            if not all(data[0:9]):
                continue
            if data[1].lower() == 'group':
                continue
            try:
                product = Menu.objects.get(item_name__iexact=data[2].strip())
                product.group = data[1].strip()
                product.price = data[3]
                product.save()
            except Menu.DoesNotExist:
                product = Menu()
                product.group = data[1].strip()
                product.item_name=data[2].strip()
                product.price = data[3]

                product.save()
        return redirect(reverse_lazy("menu_list"))
    
from django.contrib.auth.mixins import LoginRequiredMixin

# class IndexView(LoginRequiredMixin, TemplateView):
#     login_url = '/login/'
#     template_name = "scanpay/index.html"
class IndexView( TemplateView):
    login_url = '/login/'
    template_name = "scanpay/index.html"
    

# from .forms import MenuTypeForm
# class MenuTypeMixin(IsAdminMixin):
#     model = MenuType
#     form_class = MenuTypeForm
#     paginate_by = 50
#     queryset = MenuType.objects.filter(status=True, is_deleted=False, is_featured=True)
#     success_url = reverse_lazy("featured_product_list")
#     search_lookup_fields = [
#         "title",
#     ]
    
class MenuTypeProductList(ListView):
    template_name = "scanpay/menupreset/menu_preset_list.html"
    queryset = Menu.objects.filter(status=True, is_deleted=False)

    def get(self, request, *args, **kwargs):
        menu_preset_id = kwargs.get('id')
        menutype = MenuType.objects.get(id=menu_preset_id)
        menus_in_menu_preset = Menu.objects.filter(status=True, is_deleted=False, menutype=menutype)

        context = {'object_list':menus_in_menu_preset, 'menutype': menutype.title, 'menutype_id':menutype.id}

        return render(request, self.template_name, context)
        
from menu.models import MenuType
from django.views.generic.edit import CreateView
import json
from django.contrib import messages
from django.urls import reverse
class MenuTypeProductCreate(MenuTypeMixin, CreateView):
    queryset = MenuType.objects.all()
    template_name = "scanpay/menupreset/menu_preset_create.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # token = env('API_TOKEN')
        context['products'] = Menu.objects.filter(is_deleted=False, status=True)  # Replace with actual products
        # context['token'] = token
        return context


    def post(self, request, *args, **kwargs):
        menutype_id = kwargs.get('id')
        print("The menutype id is ", menutype_id)
        selected_products_json = request.POST.get('selectedProducts')
        selected_products = json.loads(selected_products_json)
        # print("I am here")
        print(selected_products_json)
        print(selected_products)

        for product in selected_products:
            selected_product_id = product['selectedProductId']
            selected_product_name = product['selectedProductName']

            # if FeaturedProducts.objects.filter(product_id=selected_product_id, is_featured=True, is_deleted=False).exists():
            #     messages.warning(request, f"Product '{selected_product_name}' is already featured.")
            # else:
            product = Menu.objects.get(id=selected_product_id)
            menutype = MenuType.objects.get(pk=menutype_id)
            product.menutype = menutype
            product.save()
        
        return redirect(reverse('menu_preset_product_list', kwargs={'id': menutype_id}))


class MenuTypeProductDelete(MenuTypeMixin, View):
    def get(self, request, *args, **kwargs):
        product_id = kwargs.get('id')
        menutype_id = kwargs.get('menutype_id')
        print(product_id)
        menu = Menu.objects.get(pk=product_id)
        menu.menutype = None
        menu.save()

        return redirect(reverse('menu_preset_product_list', kwargs={'id': menutype_id}))
        
from organization.models import Organization 
class OrganizationMixin(IsAdminMixin):
    model = Organization
    form_class = OrganizationForm
    paginate_by = 50
    queryset = Organization.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("menu_list")

from user.models import User, Customer
class OrganizationDetail(OrganizationMixin, DetailView):
    template_name = "scanpay/organization/organization_detail.html"

    def get_object(self):
        return Organization.objects.last()
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        users = User.objects.all()
        # branches = Branch.objects.filter(
        #     is_deleted=False, organization=self.get_object().id
        # )
        customers = Customer.objects.filter(is_deleted=False)
        # context["branches"] = branches
        context["customers"] = customers
        context["users"] = users

        return context

    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     users = User.objects.all()
    #     branches = Branch.objects.filter(
    #         is_deleted=False, organization=self.get_object().id
    #     )
    #     customers = Customer.objects.filter(is_deleted=False)
    #     context["branches"] = branches
    #     context["customers"] = customers
    #     context["users"] = users

    #     return context


class OrganizationCreate(OrganizationMixin, CreateView):
    template_name = "scanpay/create.html"


class OrganizationUpdate(OrganizationMixin, UpdateView):
    template_name = "scanpay/update.html"

    def get_object(self):
        return Organization.objects.last()
    

class OrganizationDelete(OrganizationMixin, DeleteMixin, View):
    pass

from .models import MediaFile
from .forms import MediaFileForm 
# from organization.models import Branch, EndDayRecord
# from .forms import BranchStockForm
class FileMixin(IsAdminMixin):
    model = MediaFile
    form_class = MediaFileForm
    paginate_by = 10
    queryset = MediaFile.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('file_list')
    search_lookup_fields = [
        "product__item_name",
        "product__description",
    ]

class FileList(FileMixin, ListView):
    template_name = "scanpay/file/file_list.html"

class FileDetail(FileMixin, DetailView):
    template_name = "file/file_detail.html"

class FileCreate(FileMixin, CreateView):
    template_name = "scanpay/create.html"

class FileUpdate(FileMixin, UpdateView):
    template_name = "scanpay/update.html"

class FileDelete(FileMixin, DeleteMixin, View):
    pass

from .forms import tbl_redeemproductForm
from .models import tbl_redeemproduct

class tbl_redeemproductMixin(IsAdminMixin):
    model = tbl_redeemproduct
    form_class = tbl_redeemproductForm
    paginate_by = 50
    queryset = tbl_redeemproduct.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("combined_view")
    search_lookup_fields = [
        "name",
        "MenuitemID__item_name",
    ]
class tbl_redeemproductList(tbl_redeemproductMixin, ListView):
    template_name = "scanpay/tbl_redeemproduct/tbl_redeemproduct-list.html"
    queryset = tbl_redeemproduct.objects.filter(status=True, is_deleted=False)


class tbl_redeemproductDetail(tbl_redeemproductMixin, DetailView):
    template_name = "scanpay/tbl_redeemproduct/tbl_redeemproduct-detail.html"


class tbl_redeemproductCreate(tbl_redeemproductMixin, CreateView):
    template_name = "scanpay/tbl_redeemproduct/tbl_redeemproduct-create.html"


class tbl_redeemproductUpdate(tbl_redeemproductMixin, UpdateView):
    template_name = "scanpay/update.html"
    

class tbl_redeemproductDelete(tbl_redeemproductMixin, DeleteMixin, View):
    pass



from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from .models import tbl_redeemproduct
from organization.models import Organization
from .forms import tbl_redeemproductForm, OrganizationForm,PointsOrganizationForm

class CombinedView(View, IsAdminMixin):
    model = tbl_redeemproduct
    form_class = tbl_redeemproductForm
    paginate_by = 50
    queryset = tbl_redeemproduct.objects.filter(status=True, is_deleted=False)
    # success_url = reverse_lazy("tbl_redeemproduct_list")
    search_lookup_fields = [
        "name",
        "MenuitemID__item_name",
    ]

    def search(self, qc):
        if self.request.GET.get("q"):
            query = self.request.GET.get("q")
            q_lookup = Q()
            for field in self.search_lookup_fields:
                q_lookup |= Q(**{field + "__icontains": query})
            return qc.filter(q_lookup)
        return qc

    def get(self, request, *args, **kwargs):
        # Handling the list view

        all_redeem_products = tbl_redeemproduct.objects.filter(status=True, is_deleted=False)
        redeem_products = self.search(all_redeem_products)
        
        # Handling the update view
        organization = Organization.objects.last()
        form = PointsOrganizationForm(instance=organization)

        return render(request, 'scanpay/combined/combined_template.html', {
            'redeem_products': redeem_products,
            'organization_form': form,
        })

    def post(self, request, *args, **kwargs):
        # Handling the form submission
        organization = Organization.objects.last()
        form = PointsOrganizationForm(request.POST, instance=organization)
        if form.is_valid():
            form.save()
            return redirect('combined_view')  # Redirect to the same view after saving
        
        # Re-render the template with errors if form is invalid
        redeem_products = tbl_redeemproduct.objects.filter(status=True, is_deleted=False)
        return render(request, 'scanpay/combined/combined_template.html', {
            'redeem_products': redeem_products,
            'organization_form': form,
        })
        
        
from django.http import JsonResponse
from .models import tbl_timedpromomenu
from .forms import TimedMenuForm 
# from organization.models import Branch, EndDayRecord
# from .forms import BranchStockForm
class TimedMenuMixin(IsAdminMixin):
    model = tbl_timedpromomenu
    form_class = TimedMenuForm
    paginate_by = 10
    queryset = tbl_timedpromomenu.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('timedmenu_list')
    search_lookup_fields = [
        "product__item_name",
        "product__description",
    ]

class TimedMenuList(TimedMenuMixin, ListView):
    template_name = "scanpay/timedmenu/timedmenu_list.html"

class TimedMenuDetail(TimedMenuMixin, DetailView):
    template_name = "scanpay/timedmenu/timedmenu_detail.html"

class TimedMenuCreate(TimedMenuMixin, CreateView):
    template_name = "scanpay/timedmenu/timedmenu_create.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add the menu types to the context
        context['menu_types'] = MenuType.objects.filter(is_timed=True)
        context['days_of_week'] = [
            ('monday', 'Monday'),
            ('tuesday', 'Tuesday'),
            ('wednesday', 'Wednesday'),
            ('thursday', 'Thursday'),
            ('friday', 'Friday'),
            ('saturday', 'Saturday'),
            ('sunday', 'Sunday')
        ]
        return context
        
class TimedMenuUpdate(TimedMenuMixin, UpdateView):
    template_name = "scanpay/update.html"

class TimedMenuDelete(TimedMenuMixin, DeleteMixin, View):
    pass